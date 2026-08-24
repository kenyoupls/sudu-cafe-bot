"""
P&L Report Generator — SUDU BY FRIENDS format
Generates a Trading Profit & Loss Account XLSX matching the café's
accounting template, auto-populated from Google Sheets data.

Categories the bot CAN auto-fill from its data:
  - Sales (from Daily Sales tab)
  - Discount allowed (from Daily Sales discount column)
  - Purchase-Food / Purchase-Beverages / Purchase-Others (from Transactions)
  - Staff Meal (if logged)
  - Waste Food (if logged)

Everything else (rent, TNB, salaries, etc.) is left as 0 for manual entry.
"""

import logging
from datetime import date, datetime
from typing import Optional
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Expense category mapping ────────────────────────────
# Maps keywords in supplier/items to P&L categories
FOOD_KEYWORDS = [
    "food", "meat", "chicken", "ayam", "fish", "ikan", "vegetable", "sayur",
    "egg", "telur", "rice", "beras", "nasi", "flour", "tepung", "oil", "minyak",
    "sauce", "sos", "sugar", "gula", "salt", "garam", "spice", "rempah",
    "frozen", "seafood", "udang", "prawn", "beef", "daging", "pork", "lamb",
    "noodle", "mee", "pasta", "bread", "roti", "butter", "margarine",
    "cream", "cheese", "tofu", "tahu", "bean", "kacang",
    "grocery", "groceries", "wet market", "pasar",
]

BEVERAGE_KEYWORDS = [
    "drink", "beverage", "minuman", "tea", "teh", "coffee", "kopi",
    "milk", "susu", "juice", "jus", "syrup", "sirap", "soda", "water",
    "ice", "ais", "bingsu", "chocolate", "cocoa", "matcha", "powder",
    "condensed", "evaporated", "creamer",
]

PACKAGING_KEYWORDS = [
    "packaging", "container", "cup", "straw", "bag", "beg", "box",
    "wrapper", "lid", "spoon", "fork", "napkin", "tissue", "plastic",
    "takeaway", "tapau", "paper", "disposable",
]


def _categorize_transaction(supplier: str, items_str: str) -> str:
    """Categorize a transaction into P&L line items."""
    text = f"{supplier} {items_str}".lower()

    for kw in BEVERAGE_KEYWORDS:
        if kw in text:
            return "beverages"

    for kw in FOOD_KEYWORDS:
        if kw in text:
            return "food"

    for kw in PACKAGING_KEYWORDS:
        if kw in text:
            return "others"

    # Default: food (most café purchases are food)
    return "food"


def generate_pnl_xlsx(month: str = None, output_path: str = None) -> Optional[str]:
    """
    Generate a P&L XLSX in SUDU BY FRIENDS format.

    Args:
        month: YYYY-MM format (default: current month)
        output_path: Where to save (default: /tmp/SUDU_PNL_{month}.xlsx)

    Returns:
        File path of generated XLSX, or None on failure.
    """
    if month is None:
        month = date.today().strftime("%Y-%m")

    if output_path is None:
        output_path = f"/tmp/SUDU_PNL_{month}.xlsx"

    try:
        # Pull data from Google Sheets
        from google_integration import (
            get_monthly_transactions,
            get_daily_sales_for_month,
        )

        transactions = get_monthly_transactions(month)
        daily_sales = get_daily_sales_for_month(month)

        # ─── Aggregate data ─────────────────────────
        # Sales
        total_sales = 0
        total_discount = 0
        total_bills = 0
        total_pax = 0

        for ds in daily_sales:
            try:
                total_sales += float(ds.get("Total Sales (RM)", 0) or 0)
                total_discount += float(ds.get("Discount (RM)", 0) or 0)
            except (ValueError, TypeError):
                pass

        # Categorize expenses
        purchase_food = 0
        purchase_beverages = 0
        purchase_others = 0
        waste_food = 0
        staff_meal = 0

        for t in transactions:
            t_type = t.get("Type", "").lower()
            if t_type not in ("expense", "purchase", "invoice", "receipt"):
                continue

            try:
                amount = float(t.get("Total (RM)", 0) or 0)
            except (ValueError, TypeError):
                amount = 0

            supplier = t.get("Supplier", "")
            items_str = t.get("Items", "")
            notes = t.get("Notes", "").lower()

            # Check for special categories
            if "waste" in notes or "waste" in items_str.lower():
                waste_food += amount
            elif "staff meal" in notes or "staff meal" in items_str.lower():
                staff_meal += amount
            else:
                cat = _categorize_transaction(supplier, items_str)
                if cat == "food":
                    purchase_food += amount
                elif cat == "beverages":
                    purchase_beverages += amount
                else:
                    purchase_others += amount

        # ─── Build workbook ─────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&L"

        # Column widths
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 10

        # Styles
        title_font = Font(name="Arial", size=14, bold=True)
        subtitle_font = Font(name="Arial", size=11, bold=True)
        header_font = Font(name="Arial", size=10, bold=True)
        section_font = Font(name="Arial", size=10, bold=True, underline="single")
        normal_font = Font(name="Arial", size=10)
        total_font = Font(name="Arial", size=10, bold=True)
        blue_font = Font(name="Arial", size=10, color="0000FF")  # manual input
        pct_font = Font(name="Arial", size=10)

        thin_border = Border(
            bottom=Side(style="thin"),
        )
        double_border = Border(
            top=Side(style="double"),
            bottom=Side(style="double"),
        )

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                  fill_type="solid")

        amt_fmt = '#,##0.00;(#,##0.00);"-"'
        pct_fmt = '0.00%'

        def write_row(row, label, amount=None, is_section=False, is_total=False,
                      is_grand_total=False, is_manual=False, formula_b=None,
                      formula_c=None):
            """Helper to write a P&L row."""
            cell_a = ws.cell(row=row, column=1, value=label)
            cell_b = ws.cell(row=row, column=2)
            cell_c = ws.cell(row=row, column=3)

            if is_section:
                cell_a.font = section_font
                return
            elif is_total:
                cell_a.font = total_font
                cell_b.font = total_font
                cell_c.font = total_font
                cell_b.border = thin_border
                cell_c.border = thin_border
            elif is_grand_total:
                cell_a.font = total_font
                cell_b.font = total_font
                cell_c.font = total_font
                cell_b.border = double_border
                cell_c.border = double_border
            else:
                cell_a.font = normal_font
                cell_b.font = blue_font if is_manual else normal_font
                cell_c.font = pct_font
                if is_manual:
                    cell_b.fill = yellow_fill

            if formula_b:
                cell_b.value = formula_b
            elif amount is not None:
                cell_b.value = amount

            if formula_c:
                cell_c.value = formula_c

            cell_b.number_format = amt_fmt
            cell_c.number_format = pct_fmt
            cell_b.alignment = Alignment(horizontal="right")
            cell_c.alignment = Alignment(horizontal="right")

        # ─── Header ─────────────────────────────────
        r = 1
        ws.cell(row=r, column=1, value="SUDU BY FRIENDS").font = title_font
        r = 2
        ws.cell(row=r, column=1,
                value="Trading Profit & Loss Account").font = subtitle_font
        r = 3
        ws.cell(row=r, column=1, value="For the month of").font = normal_font
        ws.cell(row=r, column=2, value=month).font = header_font

        # Column headers
        r = 4
        ws.cell(row=r, column=2, value="RM").font = header_font
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=3, value="% Sales").font = header_font
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")

        # ─── REVENUE ────────────────────────────────
        r = 6
        write_row(r, "Revenue", is_section=True)

        r = 7
        write_row(r, "Sales", amount=total_sales, formula_c="=IFERROR(B7/B15,0)")

        r = 8
        write_row(r, "Sales-Other Income-5% service Tax", amount=0,
                  is_manual=True, formula_c="=IFERROR(B8/B15,0)")

        r = 9
        write_row(r, "Marketing", amount=0, is_manual=True)

        r = 10
        write_row(r, "Discount allowed", amount=total_discount,
                  formula_c="=IFERROR(B10/B15,0)")

        r = 12
        write_row(r, "Service charge", amount=0, is_manual=True)

        r = 13
        write_row(r, "Service tax", amount=0, is_manual=True)

        r = 14
        write_row(r, "Sales adjustment", amount=0, is_manual=True)

        r = 15
        write_row(r, "Total Sales", is_total=True,
                  formula_b="=B7-B10+B12+B13+B14-B9",
                  formula_c="=1")

        # ─── OTHER INCOME ───────────────────────────
        r = 17
        write_row(r, "Other Income", is_section=True)

        r = 18
        write_row(r, "", amount=0, is_manual=True)

        r = 19
        write_row(r, "Royalty Fee & other income", amount=0,
                  is_manual=True, formula_c="=IFERROR(B19/B20,0)")

        r = 20
        write_row(r, "", is_total=True,
                  formula_b="=SUM(B18:B19)", formula_c="=1")

        # ─── Total Revenue line ─────────────────────
        r = 22
        ws.cell(row=r, column=2, value="=B15+B20").font = total_font
        ws.cell(row=r, column=2).number_format = amt_fmt
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")

        # ─── COST OF GOODS SOLD ─────────────────────
        r = 23
        write_row(r, "Cost of Goods Sold", is_section=True)

        r = 24
        write_row(r, "Opening Stock", amount=0, is_manual=True,
                  formula_c="=IFERROR(B24/B31,0)")

        r = 25
        write_row(r, "Purchase - Food", amount=purchase_food,
                  formula_c="=IFERROR(B25/B31,0)")

        r = 26
        write_row(r, "Purchase - Beverages", amount=purchase_beverages,
                  formula_c="=IFERROR(B26/B31,0)")

        r = 27
        write_row(r, "Purchase - Others", amount=purchase_others,
                  formula_c="=IFERROR(B27/B31,0)")

        r = 28
        write_row(r, "Waste Food-raw/complete", amount=waste_food,
                  formula_c="=IFERROR(B28/B31,0)")

        r = 29
        write_row(r, "Transport cost", amount=0, is_manual=True,
                  formula_c="=IFERROR(B29/B31,0)")

        r = 30
        write_row(r, "Closing Stock", amount=0, is_manual=True,
                  formula_c="=IFERROR(-B30/B31,0)")

        r = 31
        write_row(r, "Total Cost of Goods Sold", is_total=True,
                  formula_b="=SUM(B24:B30)",
                  formula_c="=IFERROR(B31/B15,0)")

        # ─── GROSS PROFIT ───────────────────────────
        r = 33
        write_row(r, "Gross Profit/(Loss)", is_grand_total=True,
                  formula_b="=B15+B20-B31",
                  formula_c="=IFERROR(B33/B15,0)")

        # ─── OPERATING EXPENSES ─────────────────────
        r = 35
        write_row(r, "Operating Expenses", is_section=True)

        # Restaurant Expenses
        r = 36
        write_row(r, "Restaurant Expenses", is_section=True)

        r = 37
        write_row(r, "PEST CONTROL", amount=0, is_manual=True)

        r = 38
        write_row(r, "General repair,tools,equipment", amount=0, is_manual=True)

        r = 39
        write_row(r, "Upkeep of Restaurant - AIRCON SERVICE", amount=0,
                  is_manual=True)

        r = 40
        write_row(r, "", is_total=True,
                  formula_b="=SUM(B37:B39)",
                  formula_c="=IFERROR(B40/B15,0)")

        # Rental & Utilities
        r = 42
        write_row(r, "Rental & Utilities", is_section=True)

        r = 43
        write_row(r, "TNB", amount=0, is_manual=True,
                  formula_c="=IFERROR(B43/B15,0)")

        r = 44
        write_row(r, "Water CHARGE", amount=0, is_manual=True,
                  formula_c="=IFERROR(B44/B15,0)")

        r = 45
        write_row(r, "Rental - SHOP", amount=0, is_manual=True,
                  formula_c="=IFERROR(B45/B15,0)")

        r = 46
        write_row(r, "Upkeep of Hostel", amount=0, is_manual=True)

        r = 47
        write_row(r, "Rental - Office", amount=0, is_manual=True,
                  formula_c="=IFERROR(B47/B22,0)")

        r = 48
        write_row(r, "INTERNET", amount=0, is_manual=True,
                  formula_c="=IFERROR(B48/B15,0)")

        r = 49
        write_row(r, "", is_total=True,
                  formula_b="=SUM(B43:B48)",
                  formula_c="=IFERROR(B49/B15,0)")

        # Administration Expenses
        r = 52
        write_row(r, "Administration Expenses", is_section=True)

        admin_items = [
            (53, "Salaries, Wages & Allowances"),
            (54, "FOOD HANDLING COURSE"),
            (55, "RENEW LICENSE"),
            (56, "PURCHASE CLAIM"),
            (57, "Commission"),
            (58, "Menu Book"),
            (59, "Staff Benefit & Amenities"),
            (60, "Travelling-Mileage,petrol,toll & transport"),
            (61, "Uniform/T-shirt"),
            (62, "Medical fee"),
            (63, "Introduce fee/attendance allowances"),
            (64, "Insurance premium for staff"),
            (65, "EIS Contribution"),
            (66, "SOCSO"),
            (67, "EPF"),
        ]

        for row_num, label in admin_items:
            write_row(row_num, label, amount=0, is_manual=True,
                      formula_c=f"=IFERROR(B{row_num}/B15,0)")

        r = 68
        write_row(r, "Staff Meal", amount=staff_meal,
                  formula_c="=IFERROR(B68/B15,0)")

        r = 69
        write_row(r, "", is_total=True,
                  formula_b="=SUM(B53:B68)",
                  formula_c="=IFERROR(B69/B15,0)")

        # Finance Expenses
        r = 70
        write_row(r, "Finance Expenses", is_section=True)

        r = 71
        write_row(r, "Bank Charges", amount=0, is_manual=True,
                  formula_c="=IFERROR(B71/B15,0)")

        r = 72
        write_row(r, "Credit card,food panda,grab food,shopee food",
                  amount=0, is_manual=True, formula_c="=IFERROR(B72/B15,0)")

        r = 73
        write_row(r, "Insurance,stamping,stamp duty on loan",
                  amount=0, is_manual=True, formula_c="=IFERROR(B73/B15,0)")

        r = 74
        write_row(r, "Loan interest", amount=0, is_manual=True,
                  formula_c="=IFERROR(B74/B15,0)")

        r = 75
        write_row(r, "", is_total=True,
                  formula_b="=SUM(B71:B74)",
                  formula_c="=IFERROR(B75/B15,0)")

        # Others Expenses
        r = 77
        write_row(r, "Others Expenses", is_section=True)

        other_items = [
            (78, "Insurance Premium"),
            (79, "FOOD HANDLING"),
            (80, "Postage"),
            (81, "Uniform"),
            (82, "Accounting/sec/professional fee"),
            (83, "Consultancy fee"),
            (84, "Depreciation"),
            (85, "License fee/Music copyright"),
            (86, "Penalize/Penalty/Interest"),
            (87, "Legal Fee on tenancy agreement"),
            (88, "Telephone,internet,H/P Charges/Printer"),
            (89, "Office Refreshment/maint/cleaning fee"),
            (90, "Stamping/Registered fee"),
            (91, "Upkeep of M/V - repair & Service"),
            (92, "Upkeep of office equipment/office maint"),
            (93, "Upkeep of M/V - Insurance & Road tax"),
            (94, "Membership fee"),
            (95, "Marketing exp-brochure,adv,event & etc"),
            (96, "Upkeep of office equipment"),
        ]

        for row_num, label in other_items:
            write_row(row_num, label, amount=0, is_manual=True,
                      formula_c=f"=IFERROR(B{row_num}/B15,0)")

        r = 97
        write_row(r, "", is_total=True,
                  formula_b="=SUM(B78:B96)",
                  formula_c="=IFERROR(B97/B15,0)")

        # ─── TOTAL OPERATING EXPENSES ───────────────
        r = 99
        write_row(r, "Total Operating Expenses", is_total=True,
                  formula_b="=B97+B75+B69+B49+B40",
                  formula_c="=IFERROR(B99/B15,0)")

        # ─── NET PROFIT / LOSS ──────────────────────
        r = 101
        write_row(r, "Net Profit/Loss", is_grand_total=True,
                  formula_b="=B33-B99",
                  formula_c="=IFERROR(B101/B15,0)")

        # ─── SIDE TABLE: Delivery Platform Commissions ─────
        r = 103
        ws.cell(row=r, column=1,
                value="Note: Yellow cells = manual entry needed").font = Font(
                    name="Arial", size=9, italic=True, color="666666")

        r = 105
        ws.cell(row=r, column=1,
                value="Delivery Platform Commissions").font = section_font

        headers = ["Platform", "Sales (RM)", "Rate %", "Commission (RM)"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=106, column=1 + i, value=h)
            cell.font = header_font
            cell.border = thin_border

        platforms = [
            (107, "Shopee Food", 0.25),
            (108, "Credit Card", 0.025),
            (109, "Grab", 0.33),
            (110, "FoodPanda", 0.33),
        ]

        for row_num, name, rate in platforms:
            ws.cell(row=row_num, column=1, value=name).font = normal_font
            cell_b = ws.cell(row=row_num, column=2, value=0)
            cell_b.font = blue_font
            cell_b.fill = yellow_fill
            cell_b.number_format = amt_fmt
            ws.cell(row=row_num, column=3, value=rate).number_format = "0.0%"
            ws.cell(row=row_num, column=3).font = normal_font
            ws.cell(row=row_num, column=4,
                    value=f"=B{row_num}*C{row_num}").number_format = amt_fmt
            ws.cell(row=row_num, column=4).font = normal_font

        r = 111
        ws.cell(row=r, column=1, value="Total Commissions").font = total_font
        ws.cell(row=r, column=4, value="=SUM(D107:D110)").font = total_font
        ws.cell(row=r, column=4).number_format = amt_fmt
        ws.cell(row=r, column=4).border = thin_border

        # ─── Auto-fill legend ───────────────────────
        r = 113
        ws.cell(row=r, column=1,
                value="Auto-filled from bot data:").font = Font(
                    name="Arial", size=9, bold=True)
        legend = [
            "Sales, Discount → from Daily Sales tab",
            "Purchase Food/Beverages/Others → from receipt transactions",
            "Waste Food, Staff Meal → if tagged in receipt notes",
            "All formulas (totals, %, gross profit, net P/L) → auto-calculated",
        ]
        for i, item in enumerate(legend):
            ws.cell(row=114 + i, column=1, value=f"  • {item}").font = Font(
                name="Arial", size=9, color="666666")

        # Print settings
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
            fitToPage=True
        )
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "portrait"

        # Save
        wb.save(output_path)
        logger.info(f"P&L report generated: {output_path}")
        return output_path

    except ImportError as e:
        logger.error(f"Missing dependency for P&L generation: {e}")
        return None
    except Exception as e:
        logger.error(f"Failed to generate P&L report: {e}")
        return None
